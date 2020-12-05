import openpyxl as pxl

import sys, os
import re
import json
import logging
import logging.config
import coloredlogs
from dataclasses import dataclass, field
from typing import List, Tuple


@dataclass
class Requirement:
    id: str = None
    title: str = None
    level: int = None
    subsystem: str = None
    status: List[str] = None
    priority: str = None
    parent: str = None
    description: str = None
    additional_parents: List[str] = None
    children: List[str] = field(default_factory=list)


@dataclass
class Subsystem:
    keyword: str
    level: int
    sheet_name: str
    sheet: pxl.worksheet.worksheet.Worksheet
    requirements: List[str] = None
    col_dict: dict = None
    num_requirements: int = 0
    num_TBD: int = 0
    num_TBR: int = 0
    num_XYZ: int = 0


class MRReqChecker:
    def __init__(self, filepath, row_start=3, col_start=1):
        self.workbook = pxl.load_workbook(filename=filepath)
        self.level_dict = {'LS': ['MR'],
                           'L0': ['OBJ'],
                           'L1': ['MIS'],
                           'L2': ['SYS', 'MOP'],
                           'L3': ['FAC', 'OPR', 'MCS', 'DPR', 'MEC', 'SDE', 'AVI', 'SOF', 'THR', 'POW']}
        self.sheet_names = self.workbook.sheetnames
        self.row_start = row_start
        self.col_start = col_start
        self.subsystem_dict = {}
        self.requirement_dict = {}

        # Create a custom logger
        self.setup_logging()
        self.logger = logging.getLogger(__name__)

        # add a lead to all requirements
        self.lead_req = Requirement(
            id="MR", title="MoonRanger Requirement", level=-1)
        self.requirement_dict["MR"] = self.lead_req

        for sheet_name in self.sheet_names:
            for level, keys in self.level_dict.items():
                for key in keys:
                    if re.search(f"^{key}", sheet_name, re.IGNORECASE) is not None:
                        subsys = Subsystem(keyword=key,
                                           level=int(level[-1]),
                                           sheet_name=sheet_name,
                                           sheet=self.workbook[sheet_name],
                                           col_dict=self._get_col_dict(self.workbook[sheet_name]))
                        subsys.requirements = self._get_requirements_per_subsystem(subsys)
                        subsys.num_requirements = len(subsys.requirements)
                        self.subsystem_dict[key] = subsys
        self.logger.info("MRReqChecker successfully initialized.")

    def fullsweep(self):
        """Fully sweep all requirements for errors (missing parent / additional parent, missing flag in status) and calculate statitics.
        """        
        for req_id in self.requirement_dict:
            if req_id != self.lead_req.id: 
                self.check_parent(req_id)
                num_TBD, num_TBR, num_XYZ = self.check_flag(req_id)
                subsystem = self.subsystem_dict[self.requirement_dict[req_id].subsystem]
                subsystem.num_TBD += num_TBD
                subsystem.num_TBR += num_TBR
                subsystem.num_XYZ += num_XYZ


    def convert_to_tree(self) -> dict:
        """Convert the entire workbook into a nested dictionary suitable for D3 tree visualization.

        Returns:
            dict: conform with D3 flare.json format
        """
        def _fetch_children(self, req_id: str) -> dict:
            tmp_dict = {}
            req = self.requirement_dict[req_id]
            tmp_dict["name"] = f"[{req_id}] {req.title}"
            children = req.children
            if len(children) == 0:
                tmp_dict["name"] += f" | {req.description}"
                # return tmp_dict
            else:
                tmp_dict["children"] = []
                for child_id in children:
                    tmp_dict["children"].append(
                        _fetch_children(self, child_id))
            return tmp_dict
        data = _fetch_children(self, self.lead_req.id)
        with open("data.json", "w") as outfile:
            json.dump(data, outfile)
        self.logger.info("data.json for visualization successfully generated, ready to view!")
        return data


    def output_statistics(self):
        output = {}
        total_num_requirements = 0
        total_num_TBD = 0
        total_num_TBR = 0
        total_num_XYZ = 0
        for subsystem in self.subsystem_dict.values():
            tmp = {}
            tmp["num_requirements"] = subsystem.num_requirements
            total_num_requirements += subsystem.num_requirements
            tmp["num_TBD"] = subsystem.num_TBD
            total_num_TBD += subsystem.num_TBD
            tmp["num_TBR"] = subsystem.num_TBR
            total_num_TBR += subsystem.num_TBR
            tmp["num_XYZ"] = subsystem.num_XYZ
            total_num_XYZ += subsystem.num_XYZ
            output[subsystem.keyword] = tmp
        output["total_num_requirements"] = total_num_requirements
        output["total_num_TBD"] = total_num_TBD
        output["total_num_TBR"] = total_num_TBD
        output["total_num_XYZ"] = total_num_XYZ
        with open("statistics.json", "w") as outfile:
            json.dump(output, outfile)
        self.logger.info("statistics.json for calculating TBD/TBR/XYZ successfully generated, ready to view!")
    

    def check_parent(self, req_id: str):
        """Check if a requirement has missing parent / additional parent, and establish two-way binding between its parent and it.
        """
        if not self.check_existed(req_id):
            raise ValueError("Invalid requirement ID")

        if self.requirement_dict[req_id].level >= 0:
            req = self.requirement_dict[req_id]
            if req.level != -1:
                # check parent
                parent_id = req.parent
                if not self.check_existed(parent_id):
                    self.logger.error(
                        f"{req_id} has missing or deleted parent")
                else:
                    self._add_to_parent(req_id, parent_id)
                    # self.logger.info(f"{req_id} added to parent {parent_id}")

                # check additional parent
                additional_parents = req.additional_parents
                if additional_parents is not None:
                    for parent_id in additional_parents:
                        if not self.check_existed(parent_id):
                            self.logger.error(
                                f"{req_id} has missing or deleted additional parent")
                        else:
                            self._add_to_parent(req_id, parent_id)
                            # self.logger.info(
                            #     f"{req_id} added to additional parent {parent_id}")


    def check_flag(self, req_id: str) -> Tuple[int, int, int]:
        """Check if a requirement has TBD/TBR/XYZ in its description but forgot to mark with corresponding flag in its status cell.
        """
        if not self.check_existed(req_id):
            raise ValueError("Invalid requirement ID")

        req = self.requirement_dict[req_id]
        num_TBD, num_TBR, num_XYZ = 0, 0, 0
        if req.status is not None:
            num_TBD = self._find_mark_in_description(req, "TBD")
            num_TBR = self._find_mark_in_description(req, "TBR")
            num_XYZ = self._find_mark_in_description(req, "XYZ")
            # check TBD
            if num_TBD > 0 and not self._find_flag(req, "TBD"):
                self.logger.critical(f"{req_id} should have 'TBD' flag in its status cell, now missing.")
            if num_TBD == 0 and self._find_flag(req, "TBD"):
                self.logger.warning(f"{req_id} should NOT have 'TBD' flag in its status cell, please remove it.")
            # check TBR
            if num_TBR > 0 and not self._find_flag(req, "TBR"):
                self.logger.critical(f"{req_id} should have 'TBR' flag in its status cell, now missing.")
            if num_TBR == 0 and self._find_flag(req, "TBR"):
                self.logger.warning(f"{req_id} should NOT have 'TBR' flag in its status cell, please remove it.")
            # check X, Y, Z
            if num_XYZ > 0 and not self._find_flag(req, "MISSINGVALUE"):
                self.logger.critical(f"{req_id} should have 'MissingValue' flag in its status cell, now missing.")
            if num_XYZ == 0 and self._find_flag(req, "MISSINGVALUE"):
                self.logger.warning(f"{req_id} should NOT have 'MissingValue' flag in its status cell, please remove it.")
        return num_TBD, num_TBR, num_XYZ


    def check_existed(self, req_id: str) -> bool:
        """Check if a requirement exists in the entire workbook (not just current sheet).

        Args:
            req_id (str): ID of requirement, e.g. "OBJ-1"

        Returns:
            bool: True if exists. False if not.
        """
        return req_id in self.requirement_dict.keys()


    def get_sheet_by_keyword(self, keyword: str) -> [pxl.worksheet.worksheet.Worksheet, None]:
        """Retrieve sheet from workbook based on keyword, such as "SYS" (not case-sensitive).

        Args:
            keyword (str): first letters of sheetname, e.g. "SYS"

        Returns:
            [pxl.worksheet.worksheet.Worksheet, None]: retrieved sheet
        """        
        for kw in self.subsystem_dict:
            subsys = self.subsystem_dict[kw]
            if (re.search(f"^{keyword}", kw, re.IGNORECASE) is not None) or \
               (re.search(f"^{keyword}", subsys.sheet_name, re.IGNORECASE) is not None):
                return subsys.sheet
        return None


    def _find_mark_in_description(self, req: Requirement, mark: str) -> int:
        """Find a given mark in requirement description.

        Args:
            req (Requirement): requirement object to be checked, not ID.
            mark (str): "TBD" / "TBR" / "XYZ"

        Returns:
            int: total number of mark found
        """
        if mark == "XYZ":
            regex = r"\W[XYZ]\W"
        else:
            regex = mark
        return len(re.findall(regex, req.description))


    def _find_flag(self, req: Requirement, flag: str = "Normal") -> bool:
        """Find status in a requirement's current "status" cell.

        Args:
            req (Requirement): requirement object to be checked, not ID.
            status (str, optional): status to be found, not case-sensitive. Defaults to "Normal".

        Returns:
            bool: True if find such status. False if not.
        """

        for current_status in req.status:
            if re.search(flag, current_status, re.IGNORECASE) is not None:
                return True
        return False


    def _get_requirements_per_subsystem(self, subsystem: Subsystem) -> List[Requirement]:
        """Retrieve all requirements under a certain sheet (subsystem).

        Args:
            subsystem (Subsystem): the subsystem to retrieve from

        Returns:
            List[Requirement]: a list of requirement ID.
        """        
        sheet = subsystem.sheet
        col_dict = subsystem.col_dict
        requirements = []
        # useful attributes to extract from the .xlsx file
        attrs = ['id', 'title', 'priority', 'description',
                 'status', 'parent', 'additional_parents']
        for row in sheet.iter_rows(min_row=self.row_start, max_row=sheet.max_row, values_only=True):
            if row[0] is not None and re.search("^[A-Z]{3}-", row[0], re.IGNORECASE) is not None:
                req = Requirement()
                for attr in attrs:
                    attr_idx = self._get_col_idx_by_keyword(col_dict, attr)
                    if attr_idx is not None:
                        value = row[attr_idx]
                        value = self._clean_str(value)
                    else:
                        value = None

                    if attr in ('id', 'parent'):
                        value = self._clean_str(value, handler="upper")
                        value = self._clean_id(value)

                    if attr == 'status':
                        if value is not None:
                            value = list(value.split(','))
                            value = [self._clean_str(
                                v, handler="upper") for v in value]

                    if attr == 'additional_parents':
                        if value is not None:
                            value = list(value.split(','))
                            value = [self._clean_id(self._clean_str(
                                v, handler="upper")) for v in value]

                    req.__setattr__(attr, value)

                req.level = subsystem.level
                req.subsystem = subsystem.keyword

                if req.level == 0:
                    req.parent = "MR"

                if not self._find_flag(req, flag="DELETE"):
                    self.requirement_dict[req.id] = req
                    requirements.append(req.id)
        return requirements


    def _get_col_dict(self, sheet: pxl.worksheet.worksheet.Worksheet) -> dict:
        """Retrieve column names and their indices under a certain sheet (subsystem)

        Args:
            sheet (pxl.worksheet.worksheet.Worksheet): the sheet to retrieve from

        Returns:
            dict: column name as key, column index as value.
        """        
        col_dict = {}
        counter = 0
        for col in sheet.iter_cols(1, sheet.max_column, values_only=True):
            if col[1] is not None:
                col_dict[col[1]] = counter
                counter += 1
        return col_dict


    def _get_col_idx_by_keyword(self, col_dict: dict, keyword: str) -> int:
        """Retrive column index by keyword (not case-sensitive).

        Args:
            col_dict (dict): the column dictionary to retrieve from
            keyword (str): keyword of column name

        Returns:
            int: column index
        """        
        for key in col_dict:
            if re.search(f"^{keyword}", key, re.IGNORECASE) is not None:
                return col_dict[key]
        return None

    def _add_to_parent(self, req_id: str, parent_id: str):
        """Add a requirement to its parent.

        Args:
            req_id (str): ID of requirement to add
            parent_id (str): ID of requirement to be added to
        """        
        parent = self.requirement_dict[parent_id]
        if req_id not in parent.children:
            parent.children.append(req_id)


    @staticmethod
    def _clean_str(text: [str, None], handler=None) -> str:
        if text is None:
            return None
        result = text.strip().strip('\n')
        if handler is None:
            return result
        if handler == "upper":
            return result.upper()
        elif handler == "lower":
            return result.lower()
        elif handler == "capitalize":
            return result.capitalize()
        else:
            raise ValueError("Wrong handler type to clean string")


    @staticmethod
    def _clean_id(req_id: str) -> str:
        if req_id is None or "-" not in req_id:
            return req_id
        parts = req_id.split("-")
        numeric = int(parts[-1])
        return parts[0] + "-" + str(numeric)


    @staticmethod
    def setup_logging(default_path='config.json', default_level=logging.INFO, env_key='LOG_CFG'):
        """
        | **@author:** Prathyush SP
        | Logging Setup
        """
        path = default_path
        value = os.getenv(env_key, None)
        if value:
            path = value
        if os.path.exists(path):
            try:
                config = json.loads(path)
                logging.config.dictConfig(config)
                coloredlogs.install()
            except:
                logging.basicConfig(level=default_level)
                coloredlogs.install(level=default_level)
        else:
            logging.basicConfig(level=default_level)
            coloredlogs.install(level=default_level)


if __name__ == "__main__":
    try:
        filepath = sys.argv[1]
    except:
        filepath = "MR-SYS-0001 MoonRanger Requirements.xlsx"

    rc = MRReqChecker(filepath)
    rc.fullsweep()
    # tmp = rc.requirement_dict
    # for req_id, req in tmp.items():
    #     print(req)
    rc.convert_to_tree()
    rc.output_statistics()
